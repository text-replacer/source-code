from re import I
import requests
import keyboard
import logging
from pynput import mouse
import tkinter as tk
from tkinter import messagebox
import configparser
import threading
import os
import time
import openpyxl
import atexit
import webbrowser
import sys

# Set up logging
logging.basicConfig(level=logging.INFO)

# Constants
BACKUP_XLSX_PATH = "backup_replacement_data.xlsx"
LANGUAGE_FOLDER = "languages"
DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vQKz1tiROR-S8zLK6YkrR5OPsvsuJAEVi1uC1ecTKk5-MLC-g6_jzIvSwAUNdnN5kyuIzbvU2DkmH1g/pub?output=xlsx"
DEFAULT_BEFORE_REPLACEMENT = ""
DEFAULT_AFTER_REPLACEMENT = " "
LINK_EDIT_FILE = "https://docs.google.com/spreadsheets/d/16uVFfVMKR7jVXA70g4BCo8KAE7iZVYnJT48oTpD1Z-4/edit?gid=0#gid=0"

# Global variables
last_mouse_position = (0, 0)
mouse_moved_significantly = False
is_paused = False
keyboard_thread_running = False
current_replacement_data = {}
stop_event = threading.Event()  # Global event to signal thread stop

# Load settings from .ini file
config = configparser.ConfigParser()
config.read('settings.ini')

# Load language settings
language_config = configparser.ConfigParser()

# Function to determine the base path (works in both script and frozen modes)
def get_base_path():
    if getattr(sys, 'frozen', False):
        # If the application is run as a bundle, the PyInstaller bootloader
        # extends the sys module by a flag frozen=True and sets the app
        # path into variable _MEIPASS'.
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return base_path

# Determine the path to the languages folder
BASE_PATH = get_base_path()
LANGUAGES_FOLDER = os.path.join(BASE_PATH, "languages")

# Get settings or use defaults
SHEET_URL = config.get('Settings', 'sheet_url', fallback="https://docs.google.com/spreadsheets/d/e/2PACX-1vQKz1tiROR-S8zLK6YkrR5OPsvsuJAEVi1uC1ecTKk5-MLC-g6_jzIvSwAUNdnN5kyuIzbvU2DkmH1g/pub?output=xlsx") #DEFAULT_SHEET_URL
BEFORE_REPLACEMENT = config.get('Settings', 'before_replacement', fallback=DEFAULT_BEFORE_REPLACEMENT)
AFTER_REPLACEMENT = config.get('Settings', 'after_replacement', fallback=DEFAULT_AFTER_REPLACEMENT)

# Track previous values of the widgets
previous_sheet_url = SHEET_URL
previous_before_replacement = BEFORE_REPLACEMENT
previous_after_replacement = AFTER_REPLACEMENT

def load_language(language_code):
    """
    Load the language strings from the specified language file.
    """
    global language_config

    language_file = os.path.join(LANGUAGES_FOLDER, f"{language_code}.ini")

    if os.path.exists(language_file):
        language_config.read(language_file, encoding='utf-8')
        logging.info(f"Language set to {language_code}")
    else:
        logging.error(f"Language file for {language_code} not found. Using default language.")

def download_and_process_xlsx_for_languages(xlsx_url, languages_folder):
    """
    Download the XLSX file and process it to create .ini files for each language.
    """
    try:
        response = requests.get(xlsx_url)
        response.raise_for_status()

        temp_file = os.path.join(languages_folder, "temp_languages.xlsx")
        with open(temp_file, 'wb') as f:
            f.write(response.content)

        workbook = openpyxl.load_workbook(temp_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            ini_content = []
            for row in sheet.iter_rows(values_only=True):
                if row[0]:
                    ini_content.append(row[0])

            ini_file_path = os.path.join(languages_folder, f"{sheet_name}.ini")
            with open(ini_file_path, 'w', encoding='utf-8') as ini_file:
                ini_file.write('\n'.join(ini_content))

        os.remove(temp_file)
        logging.info("Languages downloaded and processed successfully.")
    except Exception as e:
        logging.error(f"Error downloading or processing XLSX file: {e}")


def value_to_string(value):
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        else:
            return str(value)
    elif value is None:
        return ''
    else:
        return str(value)
# Rest of your code (load_replacement_data, process_xlsx_data, etc.) remains mostly the same...
# ... (Make sure to use BASE_PATH where appropriate if you have other file paths) ...
def parse_xlsx_for_replacements(workbook):
    """
    Process the XLSX data to extract replacement data, 
    including BEFORE_REPLACEMENT, AFTER_REPLACEMENT, and LINK_EDIT_FILE.
    """
    sheet = workbook.active
    replacement_data = {}
    global BEFORE_REPLACEMENT, AFTER_REPLACEMENT, LINK_EDIT_FILE  # Use global variables

    for row in sheet.iter_rows(min_row=2):
        if len(row) >= 2:
            word_cell = row[0]
            replacement_cell = row[1]
            word = value_to_string(word_cell.value).strip()  # Access .value
            replacement = value_to_string(replacement_cell.value)  # Access .value

            # Check for special cases: BEFORE_REPLACEMENT, AFTER_REPLACEMENT, and LINK_EDIT_FILE
            if word == "BEFORE_REPLACEMENT":
                BEFORE_REPLACEMENT = replacement
                logging.info(f"Set BEFORE_REPLACEMENT to: '{BEFORE_REPLACEMENT}'")
                continue
            elif word == "AFTER_REPLACEMENT":
                AFTER_REPLACEMENT = replacement
                logging.info(f"Set AFTER_REPLACEMENT to: '{AFTER_REPLACEMENT}'")
                continue
            elif word == "LINK_EDIT_FILE":
                LINK_EDIT_FILE = replacement
                logging.info(f"Set LINK_EDIT_FILE to: '{LINK_EDIT_FILE}'")
                continue

            if word and replacement is not None:
                # Replace escaped newlines (\\n) with actual newlines (\n)
                replacement = replacement.replace('\\n', '\n')
                replacement_data[word] = replacement

    return replacement_data

def load_xlsx_from_url(xlsx_url):
    """
    Load replacement data from an XLSX file hosted online.
    """
    with requests.Session() as s:
        download = s.get(xlsx_url, allow_redirects=True, stream=True, timeout=5)
        download.raise_for_status()

        # Save the downloaded XLSX to a temporary file
        temp_file = "temp_replacement_data.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(download.content)

        try:
            # Load the XLSX file using openpyxl
            workbook = openpyxl.load_workbook(temp_file)
            replacement_data = parse_xlsx_for_replacements(workbook)  # Process the workbook data

            # Remove the temporary file
            os.remove(temp_file)
            return replacement_data
        except Exception as e:
            logging.error(f"Error loading XLSX file: {e}")
            return {}

def load_xlsx_from_file(file_path):
    """
    Load replacement data from a local XLSX file.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        return parse_xlsx_for_replacements(workbook)  # Process the workbook data
    except Exception as e:
        logging.error(f"Error loading replacement data from local backup: {e}")
        return {}

def save_xlsx_to_file(replacement_data, file_path):
    """
    Save replacement data to a local XLSX file, including BEFORE_REPLACEMENT,
    AFTER_REPLACEMENT, and LINK_EDIT_FILE.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("Sheet1", 0)
    # Insert the header row
    sheet.append(["Word", "Replacement"])

    # Insert special rows
    sheet.append(["BEFORE_REPLACEMENT", BEFORE_REPLACEMENT])
    after_replacement_to_save = AFTER_REPLACEMENT.replace(' ', '\u00A0')
    sheet.append(["AFTER_REPLACEMENT", after_replacement_to_save])

    sheet.append(["LINK_EDIT_FILE", LINK_EDIT_FILE])

    # Insert the replacement data
    for word, replacement in replacement_data.items():
        sheet.append([word, replacement])

    workbook.save(file_path)

def replace_word(word, replacement):
    """
    Replace the typed word with the replacement text.
    """
    logging.info(f"Replacing '{word}' with '{replacement}'")
    # Simulate pressing backspace to delete the typed word
    for _ in range(len(word)+1):
        keyboard.press_and_release('backspace')

    # Split the replacement text by newlines and simulate typing each part
    replacement_parts = replacement.split('\n')
    keyboard.write(BEFORE_REPLACEMENT)
    for part in replacement_parts:
        keyboard.write(part)
        if part != replacement_parts[-1]:  # If not the last part, press Enter
            keyboard.press_and_release('enter')
    keyboard.write(AFTER_REPLACEMENT)
    
def on_mouse_move(x, y):
    """
    Callback function to detect significant mouse movement.
    """
    global last_mouse_position, mouse_moved_significantly
    # Check if the mouse has moved more than 10 pixels in any direction
    if abs(x - last_mouse_position[0]) > 10 or abs(y - last_mouse_position[1]) > 10:
        mouse_moved_significantly = True
        last_mouse_position = (x, y)

# Start listening for mouse events
mouse_listener = mouse.Listener(on_move=on_mouse_move)
mouse_listener.start()

def on_key_event(replacement_data):
    """
    Callback function to handle key events and detect words between spaces.
    """
    # Buffer to store typed characters
    buffer = ""

    def handle_key(event):
        nonlocal buffer
        global mouse_moved_significantly, is_paused  # Include is_paused here

        if is_paused:  # If paused, ignore all keyboard events
            return

        if event.event_type == keyboard.KEY_DOWN:
            # Reset buffer if mouse moved significantly
            if mouse_moved_significantly:
                buffer = ""
                mouse_moved_significantly = False  # Reset the flag

            if event.name == 'space':  # When a space is detected
                if buffer in replacement_data:  # Check if the word is in the replacement data
                    replace_word(buffer, replacement_data[buffer])
                buffer = ""  # Reset the buffer
            elif event.name == 'backspace':  # Handle backspace to remove the last character
                buffer = buffer[:-1]  # Remove the last character from the buffer
            elif event.name.isalnum() and event.name not in ['ctrl', 'enter']:  # Only process alphanumeric characters and ignore special keys
                buffer += event.name  # Add the character to the buffer

    return handle_key

def save_settings():
    """
    Save settings to the .ini file if the URL has changed.
    """
    global previous_sheet_url

    # Get current values from the widgets
    current_sheet_url = sheet_url_text.get("1.0", "end-1c").strip()

    # Check if the URL has changed
    if current_sheet_url != previous_sheet_url:
        # Update the previous URL value
        previous_sheet_url = current_sheet_url

        # Save the URL to the .ini file
        config['Settings'] = {
            'sheet_url': current_sheet_url
        }
        with open('settings.ini', 'w') as configfile:
            config.write(configfile)
        logging.info("Settings saved successfully!")

def load_settings_and_data():
    """
    Loads settings from the .ini file and reloads the replacement data.
    """
    global SHEET_URL, BEFORE_REPLACEMENT, AFTER_REPLACEMENT, current_replacement_data, LINK_EDIT_FILE

    # Load settings from the .ini file
    config.read('settings.ini')
    SHEET_URL = config.get('Settings', 'sheet_url', fallback=DEFAULT_SHEET_URL)

    # Reload the replacement data from the XLSX URL
    replacement_data = load_replacement_data(SHEET_URL)
    if replacement_data:
        current_replacement_data = replacement_data  # Update the current replacement data
        logging.info("Replacement data reloaded successfully.")

        # Update the Tkinter entry fields with the dynamic values
        before_replacement_entry.config(state="normal")  # Temporarily enable to update
        before_replacement_entry.delete(0, tk.END)  # Clear the entry field
        before_replacement_entry.insert(0, BEFORE_REPLACEMENT)  # Insert the dynamic value
        before_replacement_entry.config(state="readonly")  # Disable editing again

        after_replacement_entry.config(state="normal")  # Temporarily enable to update
        after_replacement_entry.delete(0, tk.END)  # Clear the entry field
        after_replacement_entry.insert(0, AFTER_REPLACEMENT)  # Insert the dynamic value
        after_replacement_entry.config(state="readonly")  # Disable editing again

        update_link_edit_file_field()  # Update the "Link Edit File" field
        return replacement_data
    else:
        logging.error("Failed to reload replacement data.")
        return None

def start_keyboard_listener(replacement_data):
    """
    Starts the keyboard hook in a separate thread.
    """
    global keyboard_thread_running
    if keyboard_thread_running:
        stop_keyboard_hook()
    stop_event.clear()  # Clear the stop event before starting the thread
    keyboard_thread = threading.Thread(target=start_keyboard_hook, args=(replacement_data,))
    keyboard_thread.daemon = True  # Daemon thread will exit when the main program exits
    keyboard_thread.start()
    keyboard_thread_running = True  # Set the flag to indicate the thread is running
    logging.info("Listening for keyboard input...")

def start_program():
    """
    Restart the program by reloading all settings and replacement data.
    """
    replacement_data = load_settings_and_data()
    if replacement_data:
        start_keyboard_listener(replacement_data)
    else:
        logging.error("Failed to start program. Exiting program.")

def start_keyboard_hook(replacement_data):
    """
    Start the keyboard hook in a separate thread.
    """
    global stop_event
    keyboard.hook(on_key_event(replacement_data))
    
    while not stop_event.is_set():  # Loop until the stop event is set
        time.sleep(0.1)  # Sleep for 100 milliseconds to reduce CPU usage
    
    # Clean up when the thread is stopped
    keyboard.unhook_all()
    logging.info("Keyboard hook stopped.")

def stop_keyboard_hook():
    """
    Signal the keyboard hook thread to stop.
    """
    global stop_event
    stop_event.set()  # Set the stop event to signal the thread to stop
    # logging.info("Stopping keyboard hook thread...")

def toggle_pause():
    """
    Toggle the pause state.
    """
    global is_paused
    is_paused = not is_paused  # Toggle the pause state
    if is_paused:
        pause_button.config(text=language_config['Buttons']['resume'], bg="red", activebackground="darkred")
        logging.info("Program paused.")
    else:
        pause_button.config(text=language_config['Buttons']['pause'], bg="green", activebackground="darkgreen")
        logging.info("Program resumed.")

def reload_xlsx_from_internet(current_value):
    """
    Reload the XLSX file from the internet and update the replacement data.
    """
    global SHEET_URL, current_replacement_data
    replacement_data = load_replacement_data(current_value)
    if replacement_data:
        current_replacement_data = replacement_data  # Update the current replacement data
        logging.info("XLSX file reloaded from the internet.")
        messagebox.showinfo("Reload", "XLSX file reloaded successfully!")
    else:
        logging.error("Failed to reload XLSX file from the internet.")
        messagebox.showerror("Reload", "Failed to reload XLSX file from the internet.")

def update_gui_language():
    """
    Update the GUI labels, buttons, and messages based on the selected language.
    """
    # Update Labels
    sheet_url_label.config(text=language_config['Labels']['sheet_url'])
    before_replacement_label.config(text=language_config['Labels']['before_replacement'])
    after_replacement_label.config(text=language_config['Labels']['after_replacement'])

    # Update Buttons
    save_button.config(text=language_config['Buttons']['save_settings'])
    pause_button.config(text=language_config['Buttons']['pause'])
    reload_button.config(text=language_config['Buttons']['reload_csv'])

def change_language(language_code):
    """
    Change the language of the program and save the selected language to the settings file.
    """
    global language_config
    load_language(language_code)
    update_gui_language()

    # Save the selected language to the settings file
    config['Settings'] = {
        'language': language_code,
        'sheet_url': sheet_url_text.get("1.0", "end-1c")
    }
    with open('settings.ini', 'w') as configfile:
        config.write(configfile)
    logging.info(f"Language changed to {language_code} and saved to settings file.")

# Button to open the Google Sheet URL
def open_google_sheet():
    url = link_edit_file_text.get("1.0", "end-1c").strip()
    if url:
        webbrowser.open(url)
    else:
        messagebox.showwarning("No URL", "The 'Link Edit File' URL is not set.")

def update_link_edit_file_field():
    link_edit_file_text.config(state="normal")  # Temporarily enable to update
    link_edit_file_text.delete("1.0", tk.END)  # Clear the entry field
    link_edit_file_text.insert("1.0", LINK_EDIT_FILE)  # Insert the dynamic value
    link_edit_file_text.config(state="disabled")  # Disable editing again

# Create the GUI
root = tk.Tk()
root.title("Text Replacer by drquochoai")
root.geometry("800x400")  # Set a fixed size for the window
root.configure(bg="#f5f5f5")  # Light background color

# Custom font
custom_font = ("Roboto", 10)

# Labels
sheet_url_label = tk.Label(root, text="Sheet URL:", font=custom_font, bg="#f5f5f5", fg="#333")
sheet_url_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")

before_replacement_label = tk.Label(root, text="Before Replacement:", font=custom_font, bg="#f5f5f5", fg="#333")
before_replacement_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")

after_replacement_label = tk.Label(root, text="After Replacement:", font=custom_font, bg="#f5f5f5", fg="#333")
after_replacement_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")

# Entry fields
# Use a Text widget for the sheet URL
sheet_url_text = tk.Text(root, width=50, height=3, font=custom_font, bg="#fff", fg="#333", relief="flat", bd=2)
sheet_url_text.insert("1.0", SHEET_URL)  # Insert the default URL
sheet_url_text.grid(row=0, column=1, padx=10, pady=10)

# Disable editing for BEFORE_REPLACEMENT and AFTER_REPLACEMENT fields
before_replacement_entry = tk.Entry(root, width=50, font=custom_font, bg="#fff", fg="#333", relief="flat", bd=2, state="readonly")
before_replacement_entry.insert(0, BEFORE_REPLACEMENT)
before_replacement_entry.grid(row=1, column=1, padx=10, pady=10)

after_replacement_entry = tk.Entry(root, width=50, font=custom_font, bg="#fff", fg="#333", relief="flat", bd=2, state="readonly")
after_replacement_entry.insert(0, AFTER_REPLACEMENT)
after_replacement_entry.grid(row=2, column=1, padx=10, pady=10)

# Buttons
save_button = tk.Button(root, text="Save Settings", font=custom_font, bg="#0078d7", fg="#fff", relief="flat", activebackground="#005a9e", activeforeground="#fff", command=save_settings)
save_button.grid(row=3, column=0, padx=10, pady=10, sticky="ew")

pause_button = tk.Button(root, text="Pause", font=custom_font, bg="green", fg="#fff", relief="flat", activebackground="#005a9e", activeforeground="#fff", command=toggle_pause)
pause_button.grid(row=3, column=1, padx=10, pady=10, sticky="ew")

reload_button = tk.Button(root, text="Reload XLSX from Internet", font=custom_font, bg="#0078d7", fg="#fff", relief="flat", activebackground="#005a9e", activeforeground="#fff", command=lambda: reload_xlsx_from_internet(sheet_url_text.get("1.0", "end-1c").strip()))
reload_button.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

# Language dropdown
saved_language = config.get('Settings', 'language', fallback='vi')
# Load all .ini files in the language folder
# Check if the languages folder exists, if not, create it and download the XLSX file
if not os.path.exists(LANGUAGES_FOLDER):
    os.makedirs(LANGUAGES_FOLDER)
    download_and_process_xlsx_for_languages("https://docs.google.com/spreadsheets/d/e/2PACX-1vSVoAsKwGTxQyR16vv8rLTwEx07N4OxZpK7qDql-tnb3sc3sOe6YCsJ549C3xFMNfMLO6Knn2I5By_Q/pub?output=xlsx", LANGUAGES_FOLDER)

load_language(saved_language)
language_files = [f for f in os.listdir(LANGUAGES_FOLDER) if f.endswith('.ini')]

language_codes = [os.path.splitext(f)[0] for f in language_files]  # Extract language codes

# Update the language dropdown with all available languages
language_var = tk.StringVar(root)
language_var.set(saved_language)  # Default language
language_menu = tk.OptionMenu(root, language_var, *language_codes, command=change_language)
language_menu.config(font=custom_font, bg="#0078d7", fg="#fff", relief="flat", activebackground="#005a9e", activeforeground="#fff")
language_menu.grid(row=4, column=2, padx=10, pady=10, sticky="ew")

# New text field for the "Link Edit File"
link_edit_file_label = tk.Label(root, text="Link Edit File:", font=custom_font, bg="#f5f5f5", fg="#333")
link_edit_file_label.grid(row=5, column=0, padx=10, pady=10, sticky="w")

link_edit_file_text = tk.Text(root, width=50, height=1, font=custom_font, bg="darkred", fg="white", relief="flat", bd=2, state="disabled")
link_edit_file_text.grid(row=5, column=1, padx=10, pady=10)

open_sheet_button = tk.Button(root, text="Open Google Sheet", font=custom_font, bg="darkred", fg="white", relief="flat", activebackground="#005a9e", activeforeground="#fff", command=open_google_sheet)
open_sheet_button.grid(row=5, column=2, padx=10, pady=10, sticky="ew")

# Call the function to update the GUI with the default language
change_language(saved_language)
update_gui_language()

# Bind save_settings() to the "<FocusOut>" event for the entry fields
sheet_url_text.bind("<FocusOut>", lambda event: save_settings())
# Status bar
status_bar = tk.Label(root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W, font=custom_font, bg="#f5f5f5", fg="#333")
status_bar.grid(row=6, column=0, columnspan=3, sticky="ew", padx=10, pady=10)

# Redirect logging to the status bar
class StatusBarHandler(logging.Handler):
    def emit(self, record):
        log_entry = self.format(record)
        status_bar.config(text=log_entry)
                
# Add the status bar handler to the logger
status_bar_handler = StatusBarHandler()
status_bar_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logging.getLogger().addHandler(status_bar_handler)

def load_replacement_data(xlsx_url):
    """
    Load replacement data from an XLSX file hosted online or from a local backup.
    """
    try:
        # Try to download the XLSX from the internet
        replacement_data = load_xlsx_from_url(xlsx_url)
        if replacement_data:
            # Save the downloaded XLSX to a local backup file
            save_xlsx_to_file(replacement_data, BACKUP_XLSX_PATH)
            logging.info("Replacement data loaded from the internet and saved to local backup.")
        return replacement_data

    except requests.exceptions.RequestException as e:
        logging.error(f"Error downloading data: {e}")
        logging.info("Attempting to load replacement data from local backup...")

        # If the download fails, try loading from the local backup file
        if os.path.exists(BACKUP_XLSX_PATH):
            return load_xlsx_from_file(BACKUP_XLSX_PATH)
        else:
            logging.error("No local backup file found. Exiting program.")
            return {}

start_program()
atexit.register(stop_keyboard_hook)

# Run the GUI
root.mainloop()
