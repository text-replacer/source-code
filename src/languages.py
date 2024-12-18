from src.starterConfig import *
import openpyxl
import requests
def load_language(language_code):

    """
    Load the language strings from the specified language file.
    """
    global language_config

    language_file = os.path.join(LANGUAGES_FOLDER, f"{language_code}.ini")

    if os.path.exists(language_file):
        language_config.read(language_file, encoding='utf-8')
        logging.info(f"Language set to {language_code}")
    else:
        logging.error(f"Language file for {language_code} not found. Using default language.")

def download_and_process_xlsx_for_languages(xlsx_url, languages_folder):
    """
    Download the XLSX file and process it to create .ini files for each language.
    """
    try:
        response = requests.get(xlsx_url)
        response.raise_for_status()

        temp_file = os.path.join(languages_folder, "temp_languages.xlsx")
        with open(temp_file, 'wb') as f:
            f.write(response.content)

        workbook = openpyxl.load_workbook(temp_file)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            ini_content = []
            for row in sheet.iter_rows(values_only=True):
                if row[0]:
                    ini_content.append(row[0])

            ini_file_path = os.path.join(languages_folder, f"{sheet_name}.ini")
            with open(ini_file_path, 'w', encoding='utf-8') as ini_file:
                ini_file.write('\n'.join(ini_content))

        os.remove(temp_file)
        logging.info("Languages downloaded and processed successfully.")
    except Exception as e:
        logging.error(f"Error downloading or processing XLSX file: {e}")
