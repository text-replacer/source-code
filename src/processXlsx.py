from re import I
import requests
import logging
import os
import openpyxl

from src.starterConfig import *
from src.languages import *

def value_to_string(value):
    """
    Convert various data types to their string representation.

    This function handles the conversion of different data types to strings with special
    handling for float numbers and None values.

    Args:
        value: The value to be converted to string. Can be of any type.

    Returns:
        str: The string representation of the input value.
            - For float numbers that are integers (e.g., 5.0), returns without decimal point
            - For float numbers with decimals, returns full float representation
            - For None values, returns empty string
            - For all other types, returns their string representation

    Examples:
        >>> value_to_string(5.0)
        '5'
        >>> value_to_string(5.5)
        '5.5'
        >>> value_to_string(None)
        ''
        >>> value_to_string('hello')
        'hello'
    """
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        else:
            return str(value)
    elif value is None:
        return ''
    else:
        return str(value)
# Rest of your code (load_replacement_data, process_xlsx_data, etc.) remains mostly the same...
# ... (Make sure to use BASE_PATH where appropriate if you have other file paths) ...
def parse_xlsx_for_replacements(workbook):
    """
    Process the XLSX data to extract replacement data, 
    including BEFORE_REPLACEMENT, AFTER_REPLACEMENT, and LINK_EDIT_FILE.
    """
    sheet = workbook.active
    replacement_data = {}
    global BEFORE_REPLACEMENT, AFTER_REPLACEMENT, LINK_EDIT_FILE  # Use global variables

    for row in sheet.iter_rows(min_row=2):
        if len(row) >= 2:
            word_cell = row[0]
            replacement_cell = row[1]
            word = value_to_string(word_cell.value).strip()  # Access .value
            replacement = value_to_string(replacement_cell.value)  # Access .value

            # Check for special cases: BEFORE_REPLACEMENT, AFTER_REPLACEMENT, and LINK_EDIT_FILE
            if word == "BEFORE_REPLACEMENT":
                BEFORE_REPLACEMENT = replacement
                logging.info(f"Set BEFORE_REPLACEMENT to: '{BEFORE_REPLACEMENT}'")
                continue
            elif word == "AFTER_REPLACEMENT":
                AFTER_REPLACEMENT = replacement
                logging.info(f"Set AFTER_REPLACEMENT to: '{AFTER_REPLACEMENT}'")
                continue
            elif word == "LINK_EDIT_FILE":
                LINK_EDIT_FILE = replacement
                logging.info(f"Set LINK_EDIT_FILE to: '{LINK_EDIT_FILE}'")
                continue

            if word and replacement is not None:
                # Replace escaped newlines (\\n) with actual newlines (\n)
                replacement = replacement.replace('\\n', '\n')
                replacement_data[word] = replacement

    return replacement_data

def load_xlsx_from_url(xlsx_url):
    """
    Load replacement data from an XLSX file hosted online.
    """
    with requests.Session() as s:
        download = s.get(xlsx_url, allow_redirects=True, stream=True, timeout=5)
        download.raise_for_status()

        # Save the downloaded XLSX to a temporary file
        temp_file = "temp_replacement_data.xlsx"
        with open(temp_file, 'wb') as f:
            f.write(download.content)

        try:
            # Load the XLSX file using openpyxl
            workbook = openpyxl.load_workbook(temp_file)
            replacement_data = parse_xlsx_for_replacements(workbook)  # Process the workbook data

            # Remove the temporary file
            os.remove(temp_file)
            return replacement_data
        except Exception as e:
            logging.error(f"Error loading XLSX file: {e}")
            return {}

def load_xlsx_from_file(file_path):
    """
    Load replacement data from a local XLSX file.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        return parse_xlsx_for_replacements(workbook)  # Process the workbook data
    except Exception as e:
        logging.error(f"Error loading replacement data from local backup: {e}")
        return {}

def save_xlsx_to_file(replacement_data, file_path):
    """
    Save replacement data to a local XLSX file, including BEFORE_REPLACEMENT,
    AFTER_REPLACEMENT, and LINK_EDIT_FILE.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("Sheet1", 0)
    # Insert the header row
    sheet.append(["Word", "Replacement"])

    # Insert special rows
    sheet.append(["BEFORE_REPLACEMENT", BEFORE_REPLACEMENT])
    after_replacement_to_save = AFTER_REPLACEMENT.replace(' ', '\u00A0')
    sheet.append(["AFTER_REPLACEMENT", after_replacement_to_save])

    sheet.append(["LINK_EDIT_FILE", LINK_EDIT_FILE])

    # Insert the replacement data
    for word, replacement in replacement_data.items():
        sheet.append([word, replacement])

    workbook.save(file_path)
